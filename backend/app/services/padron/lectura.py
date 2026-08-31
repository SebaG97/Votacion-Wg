"""Lectura del Excel del padron y auditoria de celdas combinadas (Mision 02).

Portado sin cambios de conducta desde `backend/scripts/explorar_padron.py`.
"""

from __future__ import annotations

import re
import zipfile
from pathlib import Path
from typing import Any, Sequence

from openpyxl import load_workbook

from app.services.padron.normalizacion import texto


def leer_hoja(ruta: Path, hoja: str, max_col: int) -> list[tuple[Any, ...]]:
    libro = load_workbook(ruta, read_only=True, data_only=True)
    try:
        return list(libro[hoja].iter_rows(min_row=1, max_col=max_col, values_only=True))
    finally:
        libro.close()


def leer_celdas_combinadas(ruta: Path) -> dict[str, list[str]]:
    """Extrae los rangos combinados leyendo el XML crudo del .xlsx.

    Abrir el libro sin `read_only` para consultar `ws.merged_cells` obliga a
    materializar las ~16.000 columnas fantasma; el XML da lo mismo al instante.
    """
    combinadas: dict[str, list[str]] = {}
    with zipfile.ZipFile(ruta) as z:
        libro_xml = z.read("xl/workbook.xml").decode("utf-8", "replace")
        nombres = re.findall(r'<sheet name="([^"]+)"[^>]*r:id="rId(\d+)"', libro_xml)
        rels = z.read("xl/_rels/workbook.xml.rels").decode("utf-8", "replace")
        destino = dict(
            re.findall(r'Id="rId(\d+)"[^>]*Target="(worksheets/sheet\d+\.xml)"', rels)
        )
        for nombre, rid in nombres:
            objetivo = destino.get(rid)
            if not objetivo:
                continue
            xml = z.read(f"xl/{objetivo}").decode("utf-8", "replace")
            combinadas[nombre] = re.findall(r'<mergeCell ref="([^"]+)"', xml)
    return combinadas


def indice_columna(letras: str) -> int:
    n = 0
    for ch in letras:
        n = n * 26 + (ord(ch) - 64)
    return n - 1


def auditar_combinadas(
    rangos: Sequence[str], filas: list[tuple[Any, ...]], primera_fila: int, max_col: int
) -> dict[str, Any]:
    """Clasifica cada rango combinado segun donde vive realmente el valor.

    `solo_encabezado` son los rangos donde openpyxl devolvera `None` en las filas
    siguientes: son los unicos que exigen propagacion explicita al importar.
    """
    resumen = {"vacio": 0, "valor_repetido": 0, "solo_encabezado": 0, "fuera_de_rango": 0}
    a_propagar: list[str] = []
    for ref in rangos:
        m = re.match(r"([A-Z]+)(\d+):([A-Z]+)(\d+)", ref)
        if not m:
            continue
        col_ini, fila_ini, _, fila_fin = m.group(1), int(m.group(2)), m.group(3), int(m.group(4))
        ci = indice_columna(col_ini)
        if ci >= max_col:
            resumen["fuera_de_rango"] += 1
            continue
        valores = []
        for f in range(fila_ini, fila_fin + 1):
            i = f - primera_fila
            valores.append(texto(filas[i][ci]) if 0 <= i < len(filas) else None)
        if all(v is None for v in valores):
            resumen["vacio"] += 1
        elif all(v == valores[0] for v in valores):
            resumen["valor_repetido"] += 1
        else:
            resumen["solo_encabezado"] += 1
            a_propagar.append(ref)
    return {"resumen": resumen, "rangos_a_propagar": a_propagar}
