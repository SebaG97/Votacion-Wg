"""Explorador del Excel del padron (Mision 02).

Lee `docs/Padron de ML con Jefes 2026.xlsx` en modo solo lectura, documenta la
estructura real de las tres hojas y genera los informes de calidad de datos que
alimentan `docs/PADRON_ANALISIS.md`.

El script NO modifica el Excel y NO persiste nada en base de datos: la
importacion real es la Mision 04.

Uso:

    cd backend
    python scripts/explorar_padron.py

Salidas:

    docs/padron_incidencias.csv   una fila por incidencia detectada
    docs/padron_estructura.json   estructura y metricas agregadas
"""

from __future__ import annotations

import argparse
import csv
import json
import re
import sys
import unicodedata
import zipfile
from collections import Counter, defaultdict
from dataclasses import asdict, dataclass, field
from pathlib import Path
from typing import Any, Sequence

from openpyxl import load_workbook

RAIZ = Path(__file__).resolve().parents[2]
EXCEL_POR_DEFECTO = RAIZ / "docs" / "Padron de ML con Jefes 2026.xlsx"
DIR_SALIDA_POR_DEFECTO = RAIZ / "docs"

HOJA_PRINCIPAL = "Copia de Jefes ML 2026. betty(1"
HOJA_JEFES = "LISTADO JEFES"
HOJA_PIVOT = "Hoja1"

# Ambas hojas de datos declaran un `max_column` inflado (~16.000) por formato
# residual. Los datos reales terminan en S (19) y en I (9) respectivamente, por
# eso se acota la lectura en vez de recorrer las columnas fantasma.
COLUMNAS_PRINCIPAL = 19
COLUMNAS_JEFES = 9

# Indices 0-based de la hoja principal.
P_GRUPOS = 0  # A
P_SIN_CIRCULO = 1  # B  (header con typo: "SIN CIRCILO")
P_ML = 2  # C
P_VIUDOS = 3  # D
P_MIEMBROS = 4  # E
P_CIRCULO = 5  # F
P_CIRCULO_NUEVO = 6  # G  (sin header)
P_SIN_CONSAGRACION = 7  # H
P_CONSAGRADOS = 8  # I
P_JORNADA = 9  # J
P_JEFES = 10  # K
P_MATRIMONIO = 11  # L
P_APELLIDOS = 12  # M
P_NOMBRES = 13  # N  (header es un espacio en blanco)
P_CELULAR = 14  # O
P_EMAIL = 15  # P
P_CI = 16  # Q
P_OBSERVACION = 17  # R  (sin header)
P_NO_ML = 18  # S

# Indices 0-based de LISTADO JEFES.
J_ORDEN = 0  # A (sin header)
J_CIRCULO = 1  # B
J_JEFES = 2  # C
J_MATRIMONIO = 3  # D
J_APELLIDOS = 4  # E
J_NOMBRES = 5  # F
J_CELULAR = 6  # G
J_EDUCADORES = 7  # H
J_OBSERVACION = 8  # I

SEVERIDAD_CRITICA = "CRITICA"
SEVERIDAD_ALTA = "ALTA"
SEVERIDAD_MEDIA = "MEDIA"
SEVERIDAD_BAJA = "BAJA"


# --------------------------------------------------------------------------- #
# Normalizacion
# --------------------------------------------------------------------------- #


def texto(valor: Any) -> str | None:
    """Normaliza a texto no vacio. Trata `''` y `' '` como ausencia de dato."""
    if valor is None:
        return None
    limpio = str(valor).strip()
    return limpio or None


def sin_tildes(valor: str) -> str:
    return "".join(
        c for c in unicodedata.normalize("NFD", valor) if unicodedata.category(c) != "Mn"
    )


def clave_texto(valor: Any) -> str | None:
    """Clave comparable: sin tildes, mayusculas y espacios colapsados."""
    limpio = texto(valor)
    if limpio is None:
        return None
    return re.sub(r"\s+", " ", sin_tildes(limpio)).upper()


def clave_circulo(valor: Any) -> str | None:
    """Clave de circulo: ademas quita comillas y signos de puntuacion sueltos."""
    base = clave_texto(valor)
    if base is None:
        return None
    base = base.replace('"', " ").replace("'", " ").replace("°", " ")
    return re.sub(r"\s+", " ", base).strip() or None


def normalizar_celular(valor: Any) -> tuple[str | None, str | None]:
    """Devuelve `(celular_normalizado, motivo_de_rechazo)`.

    Formato canonico paraguayo: 10 digitos con `0` inicial (`09XXXXXXXX`).
    Los valores de 9 digitos provienen de celdas cargadas como numero, que
    perdieron el cero inicial; se restaura.
    """
    crudo = texto(valor)
    if crudo is None:
        return None, "FALTANTE"
    digitos = re.sub(r"\D", "", crudo)
    if not digitos:
        return None, "SIN_DIGITOS"
    if digitos.strip("0") == "":
        # Casos `0` / `00`: placeholder de "no tiene", no un numero real.
        return None, "PLACEHOLDER_CERO"
    if len(digitos) == 9:
        digitos = "0" + digitos
    if len(digitos) != 10 or not digitos.startswith("0"):
        return None, "FORMATO_INVALIDO"
    return digitos, None


def normalizar_ci(valor: Any) -> tuple[str | None, str | None]:
    """Devuelve `(ci_normalizada, motivo_de_rechazo)`. Descarta digito verificador."""
    crudo = texto(valor)
    if crudo is None:
        return None, "FALTANTE"
    if not re.search(r"\d", crudo):
        return None, "NO_NUMERICO"
    # `1.470.451-0` -> `1470451`; `840.014` -> `840014`.
    base = crudo.split("-")[0]
    digitos = re.sub(r"\D", "", base)
    if not digitos or digitos.strip("0") == "":
        return None, "PLACEHOLDER_CERO"
    return str(int(digitos)), None


def es_marca(valor: Any) -> bool:
    """`X`, `1` y similares son marcas booleanas; `' '` y `None` no lo son."""
    return texto(valor) is not None


# --------------------------------------------------------------------------- #
# Modelo intermedio
# --------------------------------------------------------------------------- #


@dataclass
class Incidencia:
    tipo: str
    severidad: str
    hoja: str
    fila_excel: int | str
    circulo: str | None
    persona: str | None
    detalle: str

    def como_fila(self) -> dict[str, Any]:
        return asdict(self)


@dataclass
class Persona:
    fila: int
    circulo: str | None
    circulo_nuevo: str | None
    matrimonio: str | None
    apellidos: str | None
    nombres: str | None
    celular_crudo: Any
    celular: str | None
    celular_motivo: str | None
    email: str | None
    ci: str | None
    ci_motivo: str | None
    es_consagrado: bool
    es_sin_consagracion: bool
    es_ml: bool
    es_viudo: bool
    es_jefe: bool
    tiene_jornada: bool
    marca_no_ml: bool
    observacion: str | None

    @property
    def etiqueta(self) -> str:
        return " ".join(p for p in (self.apellidos, self.nombres) if p) or "(sin nombre)"


@dataclass
class Matrimonio:
    etiqueta: str | None
    circulo: str | None
    filas: list[int] = field(default_factory=list)
    personas: list[Persona] = field(default_factory=list)

    @property
    def es_consagrado(self) -> bool:
        return any(p.es_consagrado for p in self.personas)

    @property
    def es_sin_consagracion(self) -> bool:
        return any(p.es_sin_consagracion for p in self.personas)

    @property
    def es_jefe(self) -> bool:
        return any(p.es_jefe for p in self.personas)


# --------------------------------------------------------------------------- #
# Lectura
# --------------------------------------------------------------------------- #


def leer_hoja(ruta: Path, hoja: str, max_col: int) -> list[tuple[Any, ...]]:
    libro = load_workbook(ruta, read_only=True, data_only=True)
    try:
        return list(libro[hoja].iter_rows(min_row=1, max_col=max_col, values_only=True))
    finally:
        libro.close()


def leer_celdas_combinadas(ruta: Path) -> dict[str, list[str]]:
    """Extrae los rangos combinados leyendo el XML crudo del .xlsx.

    Abrir el libro sin `read_only` para consultar `ws.merged_cells` obliga a
    materializar las ~16.000 columnas fantasma; el XML da lo mismo al instante.
    """
    combinadas: dict[str, list[str]] = {}
    with zipfile.ZipFile(ruta) as z:
        libro_xml = z.read("xl/workbook.xml").decode("utf-8", "replace")
        nombres = re.findall(r'<sheet name="([^"]+)"[^>]*r:id="rId(\d+)"', libro_xml)
        rels = z.read("xl/_rels/workbook.xml.rels").decode("utf-8", "replace")
        destino = dict(
            re.findall(r'Id="rId(\d+)"[^>]*Target="(worksheets/sheet\d+\.xml)"', rels)
        )
        for nombre, rid in nombres:
            objetivo = destino.get(rid)
            if not objetivo:
                continue
            xml = z.read(f"xl/{objetivo}").decode("utf-8", "replace")
            combinadas[nombre] = re.findall(r'<mergeCell ref="([^"]+)"', xml)
    return combinadas


def indice_columna(letras: str) -> int:
    n = 0
    for ch in letras:
        n = n * 26 + (ord(ch) - 64)
    return n - 1


def auditar_combinadas(
    rangos: Sequence[str], filas: list[tuple[Any, ...]], primera_fila: int, max_col: int
) -> dict[str, Any]:
    """Clasifica cada rango combinado segun donde vive realmente el valor.

    `solo_encabezado` son los rangos donde openpyxl devolvera `None` en las filas
    siguientes: son los unicos que exigen propagacion explicita al importar.
    """
    resumen = {"vacio": 0, "valor_repetido": 0, "solo_encabezado": 0, "fuera_de_rango": 0}
    a_propagar: list[str] = []
    for ref in rangos:
        m = re.match(r"([A-Z]+)(\d+):([A-Z]+)(\d+)", ref)
        if not m:
            continue
        col_ini, fila_ini, _, fila_fin = m.group(1), int(m.group(2)), m.group(3), int(m.group(4))
        ci = indice_columna(col_ini)
        if ci >= max_col:
            resumen["fuera_de_rango"] += 1
            continue
        valores = []
        for f in range(fila_ini, fila_fin + 1):
            i = f - primera_fila
            valores.append(texto(filas[i][ci]) if 0 <= i < len(filas) else None)
        if all(v is None for v in valores):
            resumen["vacio"] += 1
        elif all(v == valores[0] for v in valores):
            resumen["valor_repetido"] += 1
        else:
            resumen["solo_encabezado"] += 1
            a_propagar.append(ref)
    return {"resumen": resumen, "rangos_a_propagar": a_propagar}


# --------------------------------------------------------------------------- #
# Clasificacion de filas de la hoja principal
# --------------------------------------------------------------------------- #


def es_fila_resumen(fila: tuple[Any, ...]) -> bool:
    """Filas de totales/porcentajes al pie de la hoja principal.

    No tienen persona ni matrimonio y acumulan agregados (o el literal
    `matrimonios` en la columna CIRCULO).
    """
    if any(texto(fila[i]) for i in (P_APELLIDOS, P_NOMBRES, P_MATRIMONIO, P_CELULAR)):
        return False
    if clave_texto(fila[P_CIRCULO]) == "MATRIMONIOS":
        return True
    for i in (P_GRUPOS, P_SIN_CIRCULO, P_ML, P_VIUDOS, P_MIEMBROS, P_SIN_CONSAGRACION, P_CONSAGRADOS):
        valor = fila[i]
        if isinstance(valor, (int, float)) and not isinstance(valor, bool) and valor > 50:
            return True
    return False


def es_fila_persona(fila: tuple[Any, ...]) -> bool:
    return bool(texto(fila[P_APELLIDOS]) or texto(fila[P_NOMBRES]))


def es_fila_encabezado_circulo(fila: tuple[Any, ...]) -> bool:
    """Fila separadora de circulo: sin persona, con conteo declarado de miembros."""
    return not es_fila_persona(fila) and texto(fila[P_MIEMBROS]) is not None


def es_fila_etiqueta(fila: tuple[Any, ...]) -> bool:
    """Fila decorativa tipo `ELLA / EL` usada como rotulo de columna."""
    valores = {clave_texto(v) for v in fila if texto(v)}
    return bool(valores) and valores <= {"EL", "ELLA", "ELLOS", "ELLAS"}


def construir_personas(filas: list[tuple[Any, ...]], desplazamiento: int) -> tuple[
    list[Persona], list[tuple[int, tuple[Any, ...]]], list[int], list[int], list[int]
]:
    personas: list[Persona] = []
    encabezados: list[tuple[int, tuple[Any, ...]]] = []
    resumen: list[int] = []
    vacias: list[int] = []
    etiquetas: list[int] = []

    for i, fila in enumerate(filas):
        nro = i + desplazamiento
        if all(texto(v) is None for v in fila):
            vacias.append(nro)
            continue
        if es_fila_resumen(fila):
            resumen.append(nro)
            continue
        if es_fila_etiqueta(fila):
            etiquetas.append(nro)
            continue
        if es_fila_persona(fila):
            cel, cel_motivo = normalizar_celular(fila[P_CELULAR])
            ci, ci_motivo = normalizar_ci(fila[P_CI])
            personas.append(
                Persona(
                    fila=nro,
                    circulo=texto(fila[P_CIRCULO]),
                    circulo_nuevo=texto(fila[P_CIRCULO_NUEVO]),
                    matrimonio=texto(fila[P_MATRIMONIO]),
                    apellidos=texto(fila[P_APELLIDOS]),
                    nombres=texto(fila[P_NOMBRES]),
                    celular_crudo=fila[P_CELULAR],
                    celular=cel,
                    celular_motivo=cel_motivo,
                    email=texto(fila[P_EMAIL]),
                    ci=ci,
                    ci_motivo=ci_motivo,
                    es_consagrado=es_marca(fila[P_CONSAGRADOS]),
                    es_sin_consagracion=es_marca(fila[P_SIN_CONSAGRACION]),
                    es_ml=es_marca(fila[P_ML]),
                    es_viudo=es_marca(fila[P_VIUDOS]),
                    es_jefe=es_marca(fila[P_JEFES]),
                    tiene_jornada=es_marca(fila[P_JORNADA]),
                    marca_no_ml=es_marca(fila[P_NO_ML]),
                    observacion=texto(fila[P_OBSERVACION]),
                )
            )
            continue
        encabezados.append((nro, fila))

    return personas, encabezados, resumen, vacias, etiquetas


def agrupar_matrimonios(personas: list[Persona]) -> list[Matrimonio]:
    """Agrupa por etiqueta MATRIMONIO + contiguidad de filas, con tope de 2.

    La etiqueta sola no alcanza: hay 7 etiquetas repetidas en circulos distintos
    (`PEREIRA FERNANDEZ`, `Reyes`, ...). La contiguidad si es fiable: en la hoja
    los conyuges estan siempre en filas consecutivas.
    """
    grupos: list[Matrimonio] = []
    actual: Matrimonio | None = None
    for p in personas:
        clave = clave_texto(p.matrimonio)
        if (
            actual is not None
            and clave is not None
            and clave_texto(actual.etiqueta) == clave
            and p.fila == actual.filas[-1] + 1
            and len(actual.filas) < 2
        ):
            actual.filas.append(p.fila)
            actual.personas.append(p)
            continue
        actual = Matrimonio(
            etiqueta=p.matrimonio,
            circulo=p.circulo,
            filas=[p.fila],
            personas=[p],
        )
        grupos.append(actual)
    return grupos


# --------------------------------------------------------------------------- #
# Deteccion de incidencias
# --------------------------------------------------------------------------- #


def detectar_incidencias(
    personas: list[Persona],
    matrimonios: list[Matrimonio],
    encabezados: list[tuple[int, tuple[Any, ...]]],
) -> list[Incidencia]:
    inc: list[Incidencia] = []

    def agregar(tipo, sev, hoja, fila, circulo, persona, detalle):
        inc.append(Incidencia(tipo, sev, hoja, fila, circulo, persona, detalle))

    # --- celular -----------------------------------------------------------
    por_celular: dict[str, list[Persona]] = defaultdict(list)
    for p in personas:
        if p.celular:
            por_celular[p.celular].append(p)
        elif p.celular_motivo == "PLACEHOLDER_CERO":
            agregar(
                "CELULAR_PLACEHOLDER", SEVERIDAD_ALTA, HOJA_PRINCIPAL, p.fila,
                p.circulo, p.etiqueta,
                f"Celular cargado como {p.celular_crudo!r}: es marcador de 'no tiene', no un numero.",
            )
        elif p.celular_motivo == "FORMATO_INVALIDO":
            agregar(
                "CELULAR_FORMATO_INVALIDO", SEVERIDAD_ALTA, HOJA_PRINCIPAL, p.fila,
                p.circulo, p.etiqueta, f"Celular no interpretable: {p.celular_crudo!r}.",
            )
        else:
            agregar(
                "CELULAR_FALTANTE", SEVERIDAD_ALTA, HOJA_PRINCIPAL, p.fila,
                p.circulo, p.etiqueta, "Persona sin celular: no podra consultar habilitacion.",
            )

    filas_matrimonio: dict[int, Matrimonio] = {}
    for m in matrimonios:
        for f in m.filas:
            filas_matrimonio[f] = m

    for cel, grupo in sorted(por_celular.items()):
        if len(grupo) < 2:
            continue
        mismo_matrimonio = len({id(filas_matrimonio.get(p.fila)) for p in grupo}) == 1
        tipo = "CELULAR_COMPARTIDO_CONYUGES" if mismo_matrimonio else "CELULAR_DUPLICADO"
        sev = SEVERIDAD_ALTA if mismo_matrimonio else SEVERIDAD_CRITICA
        detalle = "Celular repetido en filas " + ", ".join(str(p.fila) for p in grupo)
        detalle += " (" + " / ".join(p.etiqueta for p in grupo) + ")."
        if mismo_matrimonio:
            detalle += " Ambos son conyuges del mismo matrimonio."
        for p in grupo:
            agregar(tipo, sev, HOJA_PRINCIPAL, p.fila, p.circulo, p.etiqueta, f"{cel}: {detalle}")

    # --- CI ----------------------------------------------------------------
    for p in personas:
        if p.ci is None:
            sev = SEVERIDAD_BAJA if p.celular else SEVERIDAD_MEDIA
            agregar(
                "CI_FALTANTE", sev, HOJA_PRINCIPAL, p.fila, p.circulo, p.etiqueta,
                f"Sin CI utilizable (motivo: {p.ci_motivo}).",
            )
    por_ci: dict[str, list[Persona]] = defaultdict(list)
    for p in personas:
        if p.ci:
            por_ci[p.ci].append(p)
    for ci, grupo in sorted(por_ci.items()):
        if len(grupo) < 2:
            continue
        mismo_matrimonio = len({id(filas_matrimonio.get(p.fila)) for p in grupo}) == 1
        tipo = "CI_COPIADA_ENTRE_CONYUGES" if mismo_matrimonio else "CI_DUPLICADA"
        detalle = "CI repetida en filas " + ", ".join(str(p.fila) for p in grupo)
        detalle += " (" + " / ".join(p.etiqueta for p in grupo) + ")."
        if mismo_matrimonio:
            detalle += " Arrastre de la CI de un conyuge sobre el otro."
        for p in grupo:
            agregar(tipo, SEVERIDAD_MEDIA, HOJA_PRINCIPAL, p.fila, p.circulo, p.etiqueta, f"{ci}: {detalle}")

    # --- matrimonio / consagracion ----------------------------------------
    for m in matrimonios:
        if not texto(m.etiqueta):
            for p in m.personas:
                agregar(
                    "MATRIMONIO_SIN_ETIQUETA", SEVERIDAD_ALTA, HOJA_PRINCIPAL, p.fila,
                    p.circulo, p.etiqueta,
                    "Columna MATRIMONIO vacia: la persona no se puede agrupar con su conyuge.",
                )
        if len(m.personas) == 1:
            p = m.personas[0]
            if not p.es_viudo:
                agregar(
                    "MATRIMONIO_INCOMPLETO", SEVERIDAD_ALTA, HOJA_PRINCIPAL, p.fila,
                    p.circulo, p.etiqueta,
                    "Matrimonio con un solo integrante y sin marca de viudez.",
                )
        if len(m.personas) == 2:
            a, b = m.personas
            if a.es_consagrado != b.es_consagrado or a.es_sin_consagracion != b.es_sin_consagracion:
                agregar(
                    "CONSAGRACION_INCONSISTENTE", SEVERIDAD_CRITICA, HOJA_PRINCIPAL, a.fila,
                    m.circulo, f"{a.etiqueta} / {b.etiqueta}",
                    "Los dos conyuges tienen marcas de consagracion distintas: "
                    f"fila {a.fila} consagrado={a.es_consagrado} sin_consagracion={a.es_sin_consagracion}; "
                    f"fila {b.fila} consagrado={b.es_consagrado} sin_consagracion={b.es_sin_consagracion}.",
                )
            if clave_texto(a.nombres) and clave_texto(a.nombres) == clave_texto(b.nombres):
                agregar(
                    "NOMBRE_COPIADO_ENTRE_CONYUGES", SEVERIDAD_MEDIA, HOJA_PRINCIPAL, a.fila,
                    m.circulo, f"{a.etiqueta} / {b.etiqueta}",
                    f"Ambos conyuges figuran con el mismo nombre de pila ({a.nombres!r}).",
                )
        if not m.es_consagrado and not m.es_sin_consagracion:
            for p in m.personas:
                agregar(
                    "CONSAGRACION_SIN_DEFINIR", SEVERIDAD_CRITICA, HOJA_PRINCIPAL, p.fila,
                    p.circulo, p.etiqueta,
                    "Sin marca en 'Consagrados' ni en 'sin consagracion': no se puede asignar "
                    "unidad electoral.",
                )

    # --- identidad de la persona ------------------------------------------
    for p in personas:
        texto_nombre = " ".join(x for x in (p.apellidos, p.nombres) if x)
        if texto_nombre and not re.search(r"[A-Za-zÁÉÍÓÚÜÑáéíóúüñ]", texto_nombre):
            agregar(
                "NOMBRE_NO_ALFABETICO", SEVERIDAD_ALTA, HOJA_PRINCIPAL, p.fila, p.circulo,
                p.etiqueta,
                f"Las columnas de nombre contienen {texto_nombre!r}: dato cargado en la "
                "columna equivocada.",
            )

    # --- circulo -----------------------------------------------------------
    for p in personas:
        if not p.circulo:
            agregar(
                "CIRCULO_FALTANTE", SEVERIDAD_ALTA, HOJA_PRINCIPAL, p.fila, None, p.etiqueta,
                "Persona sin circulo asignado.",
            )

    etiquetas_circulo = sorted({p.circulo for p in personas if p.circulo})
    claves = {c: clave_circulo(c) for c in etiquetas_circulo}
    for a in etiquetas_circulo:
        for b in etiquetas_circulo:
            if a is b or claves[a] is None or claves[b] is None:
                continue
            if claves[a] != claves[b] and claves[b].startswith(claves[a] + " "):
                agregar(
                    "CIRCULO_ETIQUETA_VARIANTE", SEVERIDAD_MEDIA, HOJA_PRINCIPAL, "-", a, None,
                    f"{a!r} parece una escritura abreviada de {b!r}: normalizar antes de agrupar.",
                )

    # --- jefes -------------------------------------------------------------
    # Solo importa el jefe donde hay bloque no consagrado: un circulo integramente
    # consagrado vota matrimonio por matrimonio y no necesita representante.
    circulos_con_jefe = {clave_circulo(m.circulo) for m in matrimonios if m.es_jefe}
    circulos_con_bloque = {
        clave_circulo(m.circulo)
        for m in matrimonios
        if m.circulo and not m.es_consagrado
    }
    for c in etiquetas_circulo:
        if claves[c] in circulos_con_bloque and claves[c] not in circulos_con_jefe:
            agregar(
                "CIRCULO_SIN_JEFE", SEVERIDAD_CRITICA, HOJA_PRINCIPAL, "-", c, None,
                "El circulo tiene matrimonios no consagrados pero ningun integrante lleva la "
                "marca 'Jefes': el bloque no consagrado quedaria sin representante habilitado.",
            )

    for nro, fila in encabezados:
        if es_marca(fila[P_JEFES]):
            agregar(
                "JEFE_SIN_PERSONA_EN_PADRON", SEVERIDAD_CRITICA, HOJA_PRINCIPAL, nro,
                texto(fila[P_CIRCULO]), texto(fila[P_MATRIMONIO]),
                "La marca 'Jefes' esta en la fila separadora del circulo y el matrimonio jefe "
                "no tiene filas de persona (sin celular ni CI en la hoja principal).",
            )

    return inc


def reconciliar_hojas(
    personas: list[Persona], jefes_filas: list[tuple[int, tuple[Any, ...]]]
) -> tuple[list[Incidencia], dict[str, Any]]:
    inc: list[Incidencia] = []
    por_nombre: dict[tuple[str | None, str | None], list[Persona]] = defaultdict(list)
    por_celular: dict[str, list[Persona]] = defaultdict(list)
    for p in personas:
        por_nombre[(clave_texto(p.apellidos), clave_texto(p.nombres))].append(p)
        if p.celular:
            por_celular[p.celular].append(p)

    conteos = Counter()
    nombres_jefes: set[tuple[str | None, str | None]] = set()
    celulares_jefes: set[str] = set()

    # Duplicados internos de LISTADO JEFES: dos jefes distintos con el mismo
    # numero bloquearian la habilitacion por celular.
    jefes_por_celular: dict[str, list[tuple[int, str]]] = defaultdict(list)
    for nro, fila in jefes_filas:
        cel, _ = normalizar_celular(fila[J_CELULAR])
        if not cel:
            continue
        etq = " ".join(
            x for x in (texto(fila[J_APELLIDOS]), texto(fila[J_NOMBRES])) if x
        ) or "(sin nombre)"
        jefes_por_celular[cel].append((nro, etq))
    for cel, grupo in sorted(jefes_por_celular.items()):
        if len(grupo) < 2:
            continue
        for nro, etq in grupo:
            inc.append(Incidencia(
                "CELULAR_DUPLICADO_EN_LISTADO_JEFES", SEVERIDAD_CRITICA, HOJA_JEFES, nro, None, etq,
                f"{cel} figura en las filas "
                + ", ".join(f"{n} ({e})" for n, e in grupo)
                + ": no se puede resolver a que bloque habilita.",
            ))

    for nro, fila in jefes_filas:
        apellidos, nombres = texto(fila[J_APELLIDOS]), texto(fila[J_NOMBRES])
        circulo = texto(fila[J_CIRCULO])
        etiqueta = " ".join(x for x in (apellidos, nombres) if x) or "(sin nombre)"
        clave = (clave_texto(apellidos), clave_texto(nombres))
        cel, _ = normalizar_celular(fila[J_CELULAR])
        nombres_jefes.add(clave)
        if cel:
            celulares_jefes.add(cel)

        por_n = por_nombre.get(clave, [])
        por_c = por_celular.get(cel, []) if cel else []
        if por_n and por_c:
            conteos["match_nombre_y_celular"] += 1
        elif por_n:
            conteos["match_solo_nombre"] += 1
            for p in por_n:
                if p.celular and cel and p.celular != cel:
                    inc.append(Incidencia(
                        "CELULAR_DISCREPANTE_ENTRE_HOJAS", SEVERIDAD_CRITICA, HOJA_JEFES, nro,
                        circulo, etiqueta,
                        f"LISTADO JEFES tiene {fila[J_CELULAR]!r} y la hoja principal "
                        f"(fila {p.fila}) tiene {p.celular_crudo!r}.",
                    ))
        elif por_c:
            conteos["match_solo_celular"] += 1
            inc.append(Incidencia(
                "NOMBRE_DISCREPANTE_ENTRE_HOJAS", SEVERIDAD_BAJA, HOJA_JEFES, nro, circulo, etiqueta,
                "Coincide el celular con la hoja principal (fila "
                + ", ".join(str(p.fila) for p in por_c)
                + ") pero la escritura del nombre difiere.",
            ))
        else:
            conteos["sin_match"] += 1
            inc.append(Incidencia(
                "JEFE_SOLO_EN_LISTADO_JEFES", SEVERIDAD_CRITICA, HOJA_JEFES, nro, circulo, etiqueta,
                "Jefe presente en LISTADO JEFES sin correspondencia por nombre ni por celular "
                "en la hoja principal.",
            ))

    for p in personas:
        if not p.es_jefe:
            continue
        clave = (clave_texto(p.apellidos), clave_texto(p.nombres))
        if clave in nombres_jefes:
            continue
        if p.celular and p.celular in celulares_jefes:
            continue
        inc.append(Incidencia(
            "JEFE_SOLO_EN_HOJA_PRINCIPAL", SEVERIDAD_ALTA, HOJA_PRINCIPAL, p.fila, p.circulo,
            p.etiqueta,
            "Marcado como jefe en la hoja principal pero ausente de LISTADO JEFES.",
        ))

    return inc, dict(conteos)


# --------------------------------------------------------------------------- #
# Informe
# --------------------------------------------------------------------------- #


def perfilar_columnas(
    filas: list[tuple[Any, ...]], encabezado: tuple[Any, ...], max_col: int
) -> list[dict[str, Any]]:
    perfil = []
    for i in range(max_col):
        valores = [f[i] for f in filas]
        no_nulos = [v for v in valores if texto(v) is not None]
        blancos = sum(1 for v in valores if isinstance(v, str) and not v.strip())
        perfil.append({
            "columna": chr(65 + i) if i < 26 else f"col{i}",
            "encabezado": encabezado[i],
            "no_nulos": len(no_nulos),
            "solo_espacios": blancos,
            "tipos": dict(Counter(type(v).__name__ for v in no_nulos)),
            "muestra": [str(v) for v, _ in Counter(map(str, no_nulos)).most_common(5)],
        })
    return perfil


def construir_estructura(ruta: Path) -> tuple[dict[str, Any], list[Incidencia]]:
    filas_principal = leer_hoja(ruta, HOJA_PRINCIPAL, COLUMNAS_PRINCIPAL)
    filas_jefes = leer_hoja(ruta, HOJA_JEFES, COLUMNAS_JEFES)
    filas_pivot = leer_hoja(ruta, HOJA_PIVOT, 6)
    combinadas = leer_celdas_combinadas(ruta)

    enc_principal, cuerpo_principal = filas_principal[0], filas_principal[1:]
    enc_jefes, cuerpo_jefes = filas_jefes[0], filas_jefes[1:]

    personas, encabezados, resumen, vacias, etiquetas = construir_personas(cuerpo_principal, 2)
    matrimonios = agrupar_matrimonios(personas)
    jefes_personas = [
        (i + 2, f) for i, f in enumerate(cuerpo_jefes)
        if texto(f[J_APELLIDOS]) or texto(f[J_NOMBRES])
    ]

    incidencias = detectar_incidencias(personas, matrimonios, encabezados)
    inc_cruce, conteos_cruce = reconciliar_hojas(personas, jefes_personas)
    incidencias.extend(inc_cruce)

    consagrados = [m for m in matrimonios if m.es_consagrado]
    no_consagrados = [m for m in matrimonios if not m.es_consagrado and m.es_sin_consagracion]
    sin_definir = [m for m in matrimonios if not m.es_consagrado and not m.es_sin_consagracion]
    jefes_mat = [m for m in matrimonios if m.es_jefe]

    circulos = sorted({p.circulo for p in personas if p.circulo})
    circulos_con_bloque = sorted({
        clave_circulo(m.circulo) for m in no_consagrados if m.circulo
    })

    estructura = {
        "archivo": ruta.name,
        "hojas": {
            HOJA_PRINCIPAL: {
                "rol": "fuente principal de personas",
                "filas_totales": len(filas_principal),
                "columnas_reales": COLUMNAS_PRINCIPAL,
                "encabezados": [str(v) if v is not None else None for v in enc_principal],
                "filas_persona": len(personas),
                "filas_encabezado_circulo": len(encabezados),
                "filas_resumen_descartadas": resumen,
                "filas_vacias": vacias,
                "filas_etiqueta_descartadas": etiquetas,
                "perfil_columnas": perfilar_columnas(cuerpo_principal, enc_principal, COLUMNAS_PRINCIPAL),
                "celdas_combinadas": auditar_combinadas(
                    combinadas.get(HOJA_PRINCIPAL, []), cuerpo_principal, 2, COLUMNAS_PRINCIPAL
                ),
            },
            HOJA_JEFES: {
                "rol": "listado operativo de jefes/educadores por circulo",
                "filas_totales": len(filas_jefes),
                "columnas_reales": COLUMNAS_JEFES,
                "encabezados": [str(v) if v is not None else None for v in enc_jefes],
                "filas_persona": len(jefes_personas),
                "grupos_declarados": sum(1 for f in cuerpo_jefes if texto(f[J_ORDEN])),
                "perfil_columnas": perfilar_columnas(cuerpo_jefes, enc_jefes, COLUMNAS_JEFES),
                "celdas_combinadas": auditar_combinadas(
                    combinadas.get(HOJA_JEFES, []), cuerpo_jefes, 2, COLUMNAS_JEFES
                ),
            },
            HOJA_PIVOT: {
                "rol": "tabla dinamica de conteo por circulo; excluir de la importacion",
                "filas_totales": len(filas_pivot),
                "muestra": [[str(v) if v is not None else None for v in f] for f in filas_pivot[:5]],
            },
        },
        "metricas": {
            "personas": len(personas),
            "personas_con_celular": sum(1 for p in personas if p.celular),
            "personas_sin_celular": sum(1 for p in personas if not p.celular),
            "celulares_distintos": len({p.celular for p in personas if p.celular}),
            "personas_con_ci": sum(1 for p in personas if p.ci),
            "personas_sin_ci": sum(1 for p in personas if not p.ci),
            "personas_consagradas": sum(1 for p in personas if p.es_consagrado),
            "personas_sin_consagracion": sum(1 for p in personas if p.es_sin_consagracion),
            "personas_sin_marca_consagracion": sum(
                1 for p in personas if not p.es_consagrado and not p.es_sin_consagracion
            ),
            "personas_viudas": sum(1 for p in personas if p.es_viudo),
            "personas_marca_no_ml": sum(1 for p in personas if p.marca_no_ml),
            "personas_con_observacion": sum(1 for p in personas if p.observacion),
            "personas_jefe": sum(1 for p in personas if p.es_jefe),
            "matrimonios": len(matrimonios),
            "matrimonios_de_dos": sum(1 for m in matrimonios if len(m.personas) == 2),
            "matrimonios_de_uno": sum(1 for m in matrimonios if len(m.personas) == 1),
            "matrimonios_consagrados": len(consagrados),
            "matrimonios_consagrados_unipersonales": sum(
                1 for m in consagrados if len(m.personas) == 1
            ),
            "matrimonios_no_consagrados": len(no_consagrados),
            "matrimonios_sin_definir": len(sin_definir),
            "matrimonios_jefe": len(jefes_mat),
            "matrimonios_jefe_consagrados": sum(1 for m in jefes_mat if m.es_consagrado),
            "matrimonios_jefe_no_consagrados": sum(
                1 for m in jefes_mat if not m.es_consagrado
            ),
            "circulos_distintos": len(circulos),
            "circulos_con_bloque_no_consagrado": len(circulos_con_bloque),
            "circulos_con_jefe": len({clave_circulo(m.circulo) for m in jefes_mat if m.circulo}),
            "jefes_en_listado": len(jefes_personas),
            "grupos_en_listado": sum(1 for f in cuerpo_jefes if texto(f[J_ORDEN])),
            "reconciliacion_listado_jefes": conteos_cruce,
        },
        "votos_maximos_estimados": {
            "MATRIMONIO_CONSAGRADO": len(consagrados),
            "BLOQUE_NO_CONSAGRADO": len(circulos_con_bloque),
            "total": len(consagrados) + len(circulos_con_bloque),
            "nota": "Estimacion previa a resolver las incidencias criticas; no es el padron final.",
        },
        "circulos": circulos,
        "incidencias_por_tipo": dict(Counter(i.tipo for i in incidencias)),
        "incidencias_por_severidad": dict(Counter(i.severidad for i in incidencias)),
    }
    return estructura, incidencias


def escribir_salidas(
    estructura: dict[str, Any], incidencias: list[Incidencia], dir_salida: Path
) -> tuple[Path, Path]:
    dir_salida.mkdir(parents=True, exist_ok=True)
    ruta_csv = dir_salida / "padron_incidencias.csv"
    ruta_json = dir_salida / "padron_estructura.json"

    orden = {SEVERIDAD_CRITICA: 0, SEVERIDAD_ALTA: 1, SEVERIDAD_MEDIA: 2, SEVERIDAD_BAJA: 3}
    ordenadas = sorted(
        incidencias,
        key=lambda i: (orden.get(i.severidad, 9), i.tipo, str(i.fila_excel).rjust(6)),
    )
    with ruta_csv.open("w", encoding="utf-8-sig", newline="") as fh:
        writer = csv.DictWriter(
            fh,
            fieldnames=["tipo", "severidad", "hoja", "fila_excel", "circulo", "persona", "detalle"],
        )
        writer.writeheader()
        for i in ordenadas:
            writer.writerow(i.como_fila())

    with ruta_json.open("w", encoding="utf-8") as fh:
        json.dump(estructura, fh, ensure_ascii=False, indent=2)

    return ruta_csv, ruta_json


def imprimir_resumen(estructura: dict[str, Any]) -> None:
    m = estructura["metricas"]
    print(f"Archivo: {estructura['archivo']}")
    print("\n== Hojas ==")
    for nombre, datos in estructura["hojas"].items():
        print(f"  {nombre!r}: {datos['rol']} ({datos['filas_totales']} filas leidas)")
    print("\n== Personas ==")
    for clave in (
        "personas", "personas_con_celular", "personas_sin_celular", "celulares_distintos",
        "personas_sin_ci", "personas_consagradas", "personas_sin_consagracion",
        "personas_sin_marca_consagracion", "personas_viudas", "personas_jefe",
    ):
        print(f"  {clave:38s} {m[clave]}")
    print("\n== Matrimonios y grupos ==")
    for clave in (
        "matrimonios", "matrimonios_de_dos", "matrimonios_de_uno", "matrimonios_consagrados",
        "matrimonios_consagrados_unipersonales", "matrimonios_no_consagrados",
        "matrimonios_sin_definir", "matrimonios_jefe", "matrimonios_jefe_consagrados",
        "circulos_distintos", "circulos_con_jefe", "circulos_con_bloque_no_consagrado",
    ):
        print(f"  {clave:38s} {m[clave]}")
    print("\n== Votos maximos estimados ==")
    for clave, valor in estructura["votos_maximos_estimados"].items():
        print(f"  {clave:38s} {valor}")
    print("\n== Incidencias por severidad ==")
    for sev, n in sorted(estructura["incidencias_por_severidad"].items()):
        print(f"  {sev:38s} {n}")
    print("\n== Incidencias por tipo ==")
    for tipo, n in sorted(estructura["incidencias_por_tipo"].items(), key=lambda x: -x[1]):
        print(f"  {tipo:38s} {n}")


def main(argv: Sequence[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--excel", type=Path, default=EXCEL_POR_DEFECTO)
    parser.add_argument("--salida", type=Path, default=DIR_SALIDA_POR_DEFECTO)
    args = parser.parse_args(argv)

    if not args.excel.exists():
        print(f"No se encontro el Excel: {args.excel}", file=sys.stderr)
        return 1

    estructura, incidencias = construir_estructura(args.excel)
    ruta_csv, ruta_json = escribir_salidas(estructura, incidencias, args.salida)
    imprimir_resumen(estructura)
    print(f"\nIncidencias: {ruta_csv}")
    print(f"Estructura:  {ruta_json}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
