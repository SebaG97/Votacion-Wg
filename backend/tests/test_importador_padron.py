"""Pruebas del importador real del padron (Mision 04).

Usa un .xlsx sintetico chico (no el Excel real de 15 MB) que reproduce a
proposito los casos limite documentados en `docs/PADRON_ANALISIS.md`:

- Un matrimonio de un solo integrante y sin marca de viudez (`Gonzalez Pedro`,
  `Lopez Ana`): MATRIMONIO_INCOMPLETO (DEC-007).
- Dos etiquetas `MATRIMONIO` iguales ("PEREIRA FERNANDEZ") en circulos
  distintos, que no deben agruparse entre si (DEC-007).
- Un celular compartido entre conyuges (Pereira/Fernandez): no bloquea el
  voto (DEC-008).
- Una fila de resumen al pie (`CIRCULO` = "matrimonios") descartada por
  deteccion estructural (DEC-006).
- Una celda combinada (columna Jornada, sin efecto en el modelo).
- Un jefe que solo aparece en `LISTADO JEFES` sin correspondencia en la hoja
  principal (`Desconocido Nadie`): JEFE_SOLO_EN_LISTADO_JEFES (DEC-009).
- Un jefe cuyo celular falta en la hoja principal pero se completa via
  `LISTADO JEFES` por coincidencia de nombre (`Benitez Marcos`), sin que su
  matrimonio quede sin celular disponible porque su conyuge si tiene uno
  valido (DEC-009).
- Un matrimonio consagrado donde ningun integrante tiene un celular valido
  (`Diaz Roberto` / `Diaz Insfran Sonia`): MATRIMONIO_SIN_CELULAR_DISPONIBLE
  bloquea la unidad electoral (DEC-017).
- Un circulo de postulantes (`POSTULANTES B`) con jefe resuelto: debe generar
  su bloque en estado PENDIENTE_DEFINICION_POSTULANTES (DEC-013, DEC-016).

El Excel real (`docs/Padron de ML con Jefes 2026.xlsx`) solo se usa en
`test_importacion_contra_excel_real`, marcada `@pytest.mark.slow`.
"""

from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import Workbook

from app.models import Grupo, ImportacionPadron, IncidenciaPadron, Matrimonio, Persona, UnidadElectoral, Votacion
from app.models.enums import EstadoImportacion, EstadoUnidadElectoral, EstadoVotacion, TipoUnidadElectoral
from app.services.padron.columnas import HOJA_JEFES, HOJA_PRINCIPAL
from app.services.padron.importador import ImportacionRechazadaError, ejecutar_importacion

RAIZ = Path(__file__).resolve().parents[2]
EXCEL_REAL = RAIZ / "docs" / "Padron de ML con Jefes 2026.xlsx"

ENCABEZADO_PRINCIPAL = [
    "Grupos", "SIN CIRCILO", "ML", "Viudos", "miembros  x grupo / circulo",
    "CIRCULO", None, "sin consagracion", "Consagrados", "Jornada Planificacion",
    "Jefes", "MATRIMONIO", "Apellidos", " ", "CELULAR ", "E-MAIL ", "CI", None, "No ML",
]

ENCABEZADO_JEFES = [
    None, "CIRCULO", "Jefes", "MATRIMONIO", "Apellidos", "NOMBRES ", "CELULAR ", "educadores", "OBSERVACION",
]


def _construir_excel_fixture(ruta: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = HOJA_PRINCIPAL
    ws.append(ENCABEZADO_PRINCIPAL)

    filas = [
        # fila 2: separador CIRCULO A
        [1, None, None, None, 2, "CIRCULO A", None, None, None, None, None, None, None, None, None, None, None, None, None],
        # fila 3: Pereira Juan (consagrado, matrimonio PEREIRA FERNANDEZ, celular compartido)
        [None, None, 1, None, None, "CIRCULO A", None, None, 1, None, None, "PEREIRA FERNANDEZ", "Pereira", "Juan", "0981-111-111", None, "1234567", None, None],
        # fila 4: Fernandez Maria (misma pareja, mismo celular)
        [None, None, 1, None, None, "CIRCULO A", None, None, 1, None, None, "PEREIRA FERNANDEZ", "Fernandez", "Maria", "0981-111-111", None, "1234568", None, None],
        # fila 5: separador CIRCULO B
        [1, None, None, None, 1, "CIRCULO B", None, None, None, None, None, None, None, None, None, None, None, None, None],
        # fila 6: Gonzalez Pedro (etiqueta MATRIMONIO repetida, no contiguo -> grupo propio de 1,
        # sin viudez -> MATRIMONIO_INCOMPLETO; celular propio valido -> no bloquea por DEC-017)
        [None, None, 1, None, None, "CIRCULO B", None, 1, None, None, "X", "PEREIRA FERNANDEZ", "Gonzalez", "Pedro", "0982-222-222", None, "2222222", None, None],
        # fila 7: separador CIRCULO D
        [1, None, None, None, 2, "CIRCULO D", None, None, None, None, None, None, None, None, None, None, None, None, None],
        # fila 8: Diaz Roberto (consagrado, sin celular)
        [None, None, 1, None, None, "CIRCULO D", None, None, 1, None, None, "DIAZ INSFRAN", "Diaz", "Roberto", None, None, "8888881", None, None],
        # fila 9: Diaz Insfran Sonia (misma pareja, tampoco tiene celular -> MATRIMONIO_SIN_CELULAR_DISPONIBLE)
        [None, None, 1, None, None, "CIRCULO D", None, None, 1, None, None, "DIAZ INSFRAN", "Diaz Insfran", "Sonia", None, None, "8888882", None, None],
        # fila 10: separador CIRCULO E
        [1, None, None, None, 2, "CIRCULO E", None, None, None, None, None, None, None, None, None, None, None, None, None],
        # fila 11: Benitez Marcos (jefe, sin celular propio -> se completa via LISTADO JEFES)
        [None, None, 1, None, None, "CIRCULO E", None, 1, None, None, "X", "BENITEZ ROJAS", "Benitez", "Marcos", None, None, "5555551", None, None],
        # fila 12: Benitez Rojas Laura (misma pareja, celular propio valido -> el matrimonio SI tiene
        # celular disponible pese a que Marcos no tenga el suyo cargado en la hoja principal)
        [None, None, 1, None, None, "CIRCULO E", None, 1, None, None, None, "BENITEZ ROJAS", "Benitez Rojas", "Laura", "0985-555-555", None, "5555552", None, None],
        # fila 13: separador POSTULANTES B
        [1, None, None, None, 1, "POSTULANTES B", None, None, None, None, None, None, None, None, None, None, None, None, None],
        # fila 14: Lopez Ana (postulante, jefe, matrimonio de un solo integrante)
        [None, None, 1, None, None, "POSTULANTES B", None, 1, None, None, "X", "LOPEZ MARTINEZ", "Lopez", "Ana", "0984-444-444", None, "4444444", None, None],
        # fila 15: fila de resumen al pie, descartada por deteccion estructural
        [None, None, None, None, None, "matrimonios", None, None, None, None, None, None, None, None, None, None, None, None, None],
    ]
    for fila in filas:
        ws.append(fila)

    # Celda combinada inocua (columna Jornada Planificacion, sin efecto en el modelo).
    ws.merge_cells(start_row=3, start_column=10, end_row=4, end_column=10)

    ws_jefes = wb.create_sheet(HOJA_JEFES)
    ws_jefes.append(ENCABEZADO_JEFES)
    # Coincide por nombre y celular con Gonzalez Pedro (fila 6): solo confirma el jefe.
    ws_jefes.append([1, "CIRCULO B", 1, "PEREIRA FERNANDEZ", "Gonzalez", "Pedro", "0982-222-222", None, None])
    # Coincide solo por nombre con Benitez Marcos (fila 11): completa su celular faltante.
    ws_jefes.append([2, "CIRCULO E", 1, "BENITEZ ROJAS", "Benitez", "Marcos", "0983-777-777", None, None])
    # Sin correspondencia en la hoja principal ni por nombre ni por celular.
    ws_jefes.append([3, "CIRCULO Z", 1, "SIN MATCH", "Desconocido", "Nadie", "0999-999-999", None, None])

    wb.save(ruta)


@pytest.fixture()
def excel_fixture(tmp_path) -> Path:
    ruta = tmp_path / "padron_fixture.xlsx"
    _construir_excel_fixture(ruta)
    return ruta


def _grupo(db_session, nombre_normalizado: str) -> Grupo:
    grupo = (
        db_session.query(Grupo)
        .filter(Grupo.nombre_normalizado == nombre_normalizado)
        .one()
    )
    return grupo


def _unidad(db_session, tipo: TipoUnidadElectoral, referencia_id: int) -> UnidadElectoral:
    return (
        db_session.query(UnidadElectoral)
        .filter(UnidadElectoral.tipo == tipo, UnidadElectoral.referencia_id == referencia_id)
        .one()
    )


def test_importacion_crea_personas_matrimonios_y_grupos(db_session, excel_fixture):
    importacion = ejecutar_importacion(db_session, excel_fixture, usuario="tester")

    assert importacion.estado == EstadoImportacion.COMPLETADA
    assert db_session.query(Persona).count() == 8
    assert db_session.query(Matrimonio).count() == 5
    assert db_session.query(Grupo).count() == 5


def test_dos_etiquetas_matrimonio_iguales_en_circulos_distintos_no_se_fusionan(
    db_session, excel_fixture
):
    ejecutar_importacion(db_session, excel_fixture)

    matrimonios = (
        db_session.query(Matrimonio).filter(Matrimonio.codigo_externo == "PEREIRA FERNANDEZ").all()
    )
    assert len(matrimonios) == 2
    tamanos = sorted(
        1 if m.integrante_2_id is None else 2 for m in matrimonios
    )
    assert tamanos == [1, 2]


def test_celular_compartido_entre_conyuges_no_bloquea_el_voto(db_session, excel_fixture):
    ejecutar_importacion(db_session, excel_fixture)

    matrimonio_dos = (
        db_session.query(Matrimonio)
        .filter(Matrimonio.codigo_externo == "PEREIRA FERNANDEZ", Matrimonio.integrante_2_id.isnot(None))
        .one()
    )
    unidad = _unidad(db_session, TipoUnidadElectoral.MATRIMONIO_CONSAGRADO, matrimonio_dos.id)
    assert unidad.estado == EstadoUnidadElectoral.HABILITADA.value

    incidencia = (
        db_session.query(IncidenciaPadron)
        .filter(IncidenciaPadron.tipo == "CELULAR_COMPARTIDO_CONYUGES")
        .all()
    )
    assert len(incidencia) == 2


def test_matrimonio_incompleto_no_bloquea_el_bloque_no_consagrado(db_session, excel_fixture):
    ejecutar_importacion(db_session, excel_fixture)

    grupo_b = _grupo(db_session, "CIRCULO B")
    unidad = _unidad(db_session, TipoUnidadElectoral.BLOQUE_NO_CONSAGRADO, grupo_b.id)
    assert unidad.estado == EstadoUnidadElectoral.HABILITADA.value

    incidencias = (
        db_session.query(IncidenciaPadron).filter(IncidenciaPadron.tipo == "MATRIMONIO_INCOMPLETO").all()
    )
    assert len(incidencias) == 2


def test_matrimonio_sin_celular_disponible_bloquea_la_unidad(db_session, excel_fixture):
    """DEC-017: aclaracion textual del dueño del padron sobre DEC-005."""
    ejecutar_importacion(db_session, excel_fixture)

    matrimonio_diaz = (
        db_session.query(Matrimonio).filter(Matrimonio.codigo_externo == "DIAZ INSFRAN").one()
    )
    incidencias = (
        db_session.query(IncidenciaPadron)
        .filter(IncidenciaPadron.tipo == "MATRIMONIO_SIN_CELULAR_DISPONIBLE")
        .all()
    )
    assert len(incidencias) == 2
    assert all(i.severidad.value == "CRITICA" for i in incidencias)

    unidad = _unidad(db_session, TipoUnidadElectoral.MATRIMONIO_CONSAGRADO, matrimonio_diaz.id)
    assert unidad.estado == EstadoUnidadElectoral.BLOQUEADA_POR_INCIDENCIA.value


def test_matrimonio_con_un_celular_valido_no_activa_dec_017(db_session, excel_fixture):
    """Benitez Marcos no tiene celular propio en la hoja principal, pero su
    conyuge si: el matrimonio tiene celular disponible y no dispara
    MATRIMONIO_SIN_CELULAR_DISPONIBLE, aunque el bloque de su circulo dependa
    de que su celular se complete via LISTADO JEFES."""
    ejecutar_importacion(db_session, excel_fixture)

    incidencias = (
        db_session.query(IncidenciaPadron)
        .filter(
            IncidenciaPadron.tipo == "MATRIMONIO_SIN_CELULAR_DISPONIBLE",
        )
        .all()
    )
    circulos_afectados = {i.descripcion for i in incidencias}
    assert not any("CIRCULO E" in c for c in circulos_afectados)

    grupo_e = _grupo(db_session, "CIRCULO E")
    unidad = _unidad(db_session, TipoUnidadElectoral.BLOQUE_NO_CONSAGRADO, grupo_e.id)
    assert unidad.estado == EstadoUnidadElectoral.HABILITADA.value


def test_celular_de_jefe_se_completa_via_listado_jefes(db_session, excel_fixture):
    ejecutar_importacion(db_session, excel_fixture)

    persona = (
        db_session.query(Persona)
        .filter(Persona.apellidos == "Benitez", Persona.nombres == "Marcos")
        .one()
    )
    assert persona.celular == "0983777777"
    assert persona.es_jefe_grupo is True


def test_jefe_solo_en_listado_jefes_genera_incidencia_critica(db_session, excel_fixture):
    ejecutar_importacion(db_session, excel_fixture)

    incidencia = (
        db_session.query(IncidenciaPadron)
        .filter(IncidenciaPadron.tipo == "JEFE_SOLO_EN_LISTADO_JEFES")
        .one()
    )
    assert incidencia.severidad.value == "CRITICA"
    assert incidencia.persona_id is None


def test_circulo_postulantes_queda_pendiente_de_definicion(db_session, excel_fixture):
    ejecutar_importacion(db_session, excel_fixture)

    grupo_postulantes = _grupo(db_session, "POSTULANTES B")
    unidad = _unidad(db_session, TipoUnidadElectoral.BLOQUE_NO_CONSAGRADO, grupo_postulantes.id)
    assert unidad.estado == EstadoUnidadElectoral.PENDIENTE_DEFINICION_POSTULANTES.value


def test_fila_de_resumen_al_pie_no_genera_persona_ni_incidencia(db_session, excel_fixture):
    ejecutar_importacion(db_session, excel_fixture)

    assert db_session.query(Persona).filter(Persona.apellidos == "").count() == 0
    assert db_session.query(Grupo).filter(Grupo.nombre_normalizado == "MATRIMONIOS").count() == 0


def test_incidencias_quedan_vinculadas_a_la_importacion(db_session, excel_fixture):
    importacion = ejecutar_importacion(db_session, excel_fixture)

    incidencias = (
        db_session.query(IncidenciaPadron)
        .filter(IncidenciaPadron.importacion_id == importacion.id)
        .count()
    )
    assert incidencias > 0
    assert incidencias == db_session.query(IncidenciaPadron).count()


def test_resumen_incluye_votos_maximos_por_tipo(db_session, excel_fixture):
    importacion = ejecutar_importacion(db_session, excel_fixture)

    assert importacion.resumen is not None
    votos_maximos = importacion.resumen["votos_maximos"]
    # Matrimonio consagrado habilitado: solo Pereira/Fernandez (Diaz/Insfran queda
    # bloqueado por DEC-017). Bloques habilitados: CIRCULO B y CIRCULO E; el de
    # POSTULANTES B queda pendiente y no cuenta como voto maximo.
    assert votos_maximos["por_tipo"]["MATRIMONIO_CONSAGRADO"] == 1
    assert votos_maximos["por_tipo"]["BLOQUE_NO_CONSAGRADO"] == 2
    assert votos_maximos["total"] == 3


def test_reimportar_es_idempotente_mientras_la_votacion_esta_en_borrador(db_session, excel_fixture):
    ejecutar_importacion(db_session, excel_fixture)
    ejecutar_importacion(db_session, excel_fixture)

    assert db_session.query(Persona).count() == 8
    assert db_session.query(Matrimonio).count() == 5
    assert db_session.query(Grupo).count() == 5
    assert db_session.query(ImportacionPadron).count() == 2
    assert db_session.query(ImportacionPadron).filter(
        ImportacionPadron.estado == EstadoImportacion.COMPLETADA
    ).count() == 2


def test_reimportar_se_rechaza_si_hay_votacion_abierta(db_session, excel_fixture):
    ejecutar_importacion(db_session, excel_fixture)

    votacion = Votacion(nombre="Votacion En Curso", estado=EstadoVotacion.ABIERTA)
    db_session.add(votacion)
    db_session.commit()

    personas_antes = db_session.query(Persona).count()
    importaciones_antes = db_session.query(ImportacionPadron).count()

    with pytest.raises(ImportacionRechazadaError):
        ejecutar_importacion(db_session, excel_fixture)

    assert db_session.query(Persona).count() == personas_antes
    assert db_session.query(ImportacionPadron).count() == importaciones_antes


def test_excel_inexistente_lanza_file_not_found(db_session, tmp_path):
    with pytest.raises(FileNotFoundError):
        ejecutar_importacion(db_session, tmp_path / "no_existe.xlsx")


@pytest.mark.slow
def test_importacion_contra_excel_real(db_session):
    """Integracion contra el Excel real. Correr manualmente antes de una votacion real:

        cd backend
        python -m pytest tests/test_importador_padron.py -m slow
    """
    if not EXCEL_REAL.exists():
        pytest.skip(f"No se encontro el Excel real en {EXCEL_REAL}")

    importacion = ejecutar_importacion(db_session, EXCEL_REAL)

    assert importacion.estado == EstadoImportacion.COMPLETADA
    assert db_session.query(Persona).count() == 1113
    assert db_session.query(Matrimonio).count() == 571
    assert db_session.query(Grupo).count() == 93
    assert (
        db_session.query(Matrimonio).filter(Matrimonio.es_consagrado.is_(True)).count() == 260
    )
    assert importacion.resumen["grupos"]["total"] == 93
